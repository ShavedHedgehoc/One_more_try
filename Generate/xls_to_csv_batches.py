import os
import csv
import openpyxl
from openpyxl import load_workbook

if __name__ == '__main__':
    #wb = load_workbook('./Пользователи.xlsx')
    wb = load_workbook('./Варки2.xlsx')
    ws=wb['Лист1']
    report_filename = 'Batches.csv'
    with open(report_filename, 'w', encoding='utf-8', newline='') as csvfile:
        csv_file = csv.writer(csvfile, delimiter=';')
        #csv_file.writerow(["Lot","Pr_lot","Code","Name","Provider","Producer","Expire"])
        csv_file.writerow(["Batch","Code","Name","Quant"])        
        for row in ws.values:
             if row[0]!=None:
                 single_row = [
                     row[0],
                     row[1].strip(" "),
                     row[2],
                     row[3],
        ]
                 csv_file.writerow(single_row)
        csvfile.close
    print("Done!")
