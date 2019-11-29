import os
import csv
import openpyxl
from openpyxl import load_workbook

if __name__ == '__main__':
    #wb = load_workbook('./Пользователи.xlsx')
    wb = load_workbook('./Партии производителя.xlsx')
    ws=wb['Лист1']
    report_filename = 'Lotpr.csv'
    with open(report_filename, 'w', encoding='utf-8', newline='') as csvfile:
        csv_file = csv.writer(csvfile, delimiter=';')
        #csv_file.writerow(["Lot","Pr_lot","Code","Name","Provider","Producer","Expire"])
        csv_file.writerow(["Lot","Pr_lot","Name","Provider","Producer","Expire"])        
        for row in ws.values:
            lot_length=len(str(row[0]).replace(" ",""))
            lot_date=row[5]            
            if lot_length==20 and lot_date!=None:                
                csv_file.writerow(row)
        csvfile.close
    print("Done!")
